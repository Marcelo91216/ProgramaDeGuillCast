import openpyxl
from openpyxl.styles import Font
from tkinter import Tk, Frame, Label, Entry, Button, Scrollbar, Listbox, DISABLED, NORMAL, VERTICAL, END, ANCHOR, Toplevel, filedialog
from tkinter import ttk

root = Tk()
root.title("Inventary GC")
root.resizable(False, False)

def showBandas(selectedsheet):
    for i in range(selectedsheet.max_row-1):
        bandas.insert(i, ( str(i+1),'->' ,selectedsheet['B'+str(i+2)].value,selectedsheet['C'+str(i+2)].value,selectedsheet['D'+str(i+2)].value ))
def openfile():
    filepath = ''
    filepath = filedialog.askopenfilename(title="Seleccione su archivo Excel", filetypes=[("Archivos xlsx","*.xlsx")])
    if(filepath != ''):
        add.config(state=NORMAL)
        find.config(state=NORMAL)
        nspopup.config(state=NORMAL)
        sub.config(state=NORMAL)
        subAll.config(state=NORMAL)
        edit.config(state=NORMAL)
        
        book = openpyxl.load_workbook(filepath)
        selectedsheet = book.active
        selectsheet.config(values=book.sheetnames)
        
        tempTxt = filepath.split('/')
        bookinuse.config(text=tempTxt[-1])
        whichsheet.config(text= selectedsheet.title)
        allpathfile.config(text=filepath)
        
        bandas.grid(row=1, column=0, columnspan=3, rowspan=2)
        bandas.delete(0, END)
        showBandas(selectedsheet)
    
def agregar():
    book = openpyxl.load_workbook(allpathfile.cget("text"))
    if medidas.get() != '' and whichsheet.cget("text") != '---':
        sheet = book[whichsheet.cget("text")]
        # Find where to put it and place it
        actualAcell = "=A"+str(sheet.max_row)+"+1"
        sheet.append([actualAcell ,medidas.get() ,propiedades.get() ,importe.get() ])
        # Update letter style
        sheet['A'+str(sheet.max_row)].font = Font(name='Arial', size=18)
        sheet['B'+str(sheet.max_row)].font = Font(name='Arial', size=18)
        sheet['C'+str(sheet.max_row)].font = Font(name='Arial', size=18)
        sheet['D'+str(sheet.max_row)].font = Font(name='Arial', size=18)
        book.save(allpathfile.cget("text"))
        info.config(text="Datos subidos y actualizados")
        medidas.delete(0, 'end')
        propiedades.delete(0, 'end')
        importe.delete(0, 'end')
        bandas.delete(0, END)
        showBandas(sheet)
    else:
        info.config(text="Su banda almenos debe\ntener unas medidas")
    
def quitar():
    tpl = bandas.get(ANCHOR)
    if tpl != '':
        book = openpyxl.load_workbook(allpathfile.cget("text"))
        sheet = book[whichsheet.cget("text")]
        sheet.delete_rows(int(tpl[0])+1)
        book.save(allpathfile.cget("text"))
        bandas.delete(0, END)
        showBandas(sheet)
        info.config(text="Dato eliminado")
    else:
        info.config(text="Seleccione un elemento")
def editar():
    tpl = bandas.get(ANCHOR)
    book = openpyxl.load_workbook(allpathfile.cget("text"))
    sheet = book[whichsheet.cget("text")]
    if tpl != '':
        if medidas.get() != '' and propiedades.get() != '' and importe.get() != '':
            puesto = int(tpl[0])+1
            sheet['B'+str(puesto)] = medidas.get()
            sheet['C'+str(puesto)] = propiedades.get()
            sheet['D'+str(puesto)] = importe.get()
            book.save(allpathfile.cget("text"))
        else:
            info.config(text='Al editar una fila\ndebe editar todos los datos')
    else:
        info.config(text="Seleccione un elemento")
    bandas.delete(0, END)
    showBandas(sheet)
    medidas.delete(0, 'end')
    propiedades.delete(0, 'end')
    importe.delete(0, 'end')
def encontrar():
    def auxFind(tam, phrase, ele):
        if phrase != '':
            cont=0
            for i in range(tam):
                if(phrase != str(((bandas.get(cont)))[ele])[0: len(phrase)]):
                    bandas.delete(cont)
                else:
                    cont=cont+1
            info.config(text="Aqui estan tus resultados")
    if medidas.get() == '' and propiedades.get() == '' and importe.get() == '':
        info.config(text="Evite dejar algun espacio vacío")
    auxFind(bandas.size(), str(medidas.get()), 2)
    auxFind(bandas.size(), str(propiedades.get()), 3)
    auxFind(bandas.size(), str(importe.get()), 4)
        
def showPopup():
    def addSheet():
        if newsheetentry.get() != '':
            book = openpyxl.load_workbook(allpathfile.cget("text"))
            book.create_sheet(newsheetentry.get())
            book.save(allpathfile.cget("text"))
            sheet = book[newsheetentry.get()]
            sheet['B1'].value = 'MEDIDAS'
            sheet['C1'].value = 'PROPIEDADES'
            sheet['D1'].value = 'IMPORTE'
            sheet['B1'].font = Font(name='Arial', size=18)
            sheet['C1'].font = Font(name='Arial', size=18)
            sheet['D1'].font = Font(name='Arial', size=18)
            book.save(allpathfile.cget("text"))
            selectsheet.config(values=book.sheetnames)
            popup.destroy()
        else:
            newsheetlabel.config(text="Ponga un nombre\nvalido")
    def delSheet():
        if delsheetcb.current() != -1:
            whichsheet.config(text='---')
            book = openpyxl.load_workbook(allpathfile.cget("text"))
            if len(book.sheetnames) != 1:
                book.remove(book[delsheetcb.get()])
                book.save(allpathfile.cget("text"))
                selectsheet.config(values=book.sheetnames)
                bandas.delete(0, END)
                popup.destroy()
            else:
                delsheetlabel.config(text='Al menos deje \nuna hoja viva')
        else:
            delsheetlabel.config(text="Esa hoja no existe")
    # New sheet
    popup = Toplevel(root)
    popup.title("Acciones de hoja")
    popup.resizable(False, False)
    
    newsheetlabel = Label(popup, text="Nombre de\nla nueva hoja")
    newsheetlabel.grid(row=0, column=0)
    newsheetentry = Entry(popup)
    newsheetentry.grid(row=1, column=0)
    newsheet = Button(popup, text="Crear", width=12, height=2, command=addSheet)
    newsheet.grid(row=2, column=0)
    # Eliminate sheet
    delsheetlabel = Label(popup, text="Que hoja \ndesea eliminar")
    delsheetlabel.grid(row=0, column=1)
    delsheetcb = ttk.Combobox(popup, width=10, height=2)
    delsheetcb.grid(row=1, column=1)
    book = openpyxl.load_workbook(allpathfile.cget("text"))
    delsheetcb.config(values=book.sheetnames)
    delsheetbtn = Button(popup, text="Borrar", width=10, height=2, command=delSheet)
    delsheetbtn.grid(row=2, column=1)
    
def changeSheet():
    if selectsheet.current() != -1:
        info.config(text="Aqui se mostraran los datos buscados, agregados y quitados")
        whichsheet.config(text=selectsheet.get())
        book = openpyxl.load_workbook(allpathfile.cget("text"))
        sheet = book[selectsheet.get()]
        bandas.delete(0, END)
        showBandas(sheet)
    else:
        bandas.delete(0, END)
        info.config(text="Use una hoja valida")
        whichsheet.config(text="---")
def erasePopup():
    def eraseAll():
        if bandas.size() > 0:
            book = openpyxl.load_workbook(allpathfile.cget("text"), data_only = True)
            sheet = book[whichsheet.cget("text")]
            cont = -1
            for i in range(bandas.size()):
                #print('A' , sheet['A'+str(int((bandas.get(i))[0])-cont)].value, '->', str(int((bandas.get(i))[0])-cont))
                sheet.delete_rows(int((bandas.get(i))[0])-cont)
                cont=cont+1
            book.save(allpathfile.cget("text"))
            bandas.delete(0, END)
            showBandas(sheet)
            delPop.destroy()
        else:
            eraseAllLbl.config(text="No hay nada")
    delPop = Toplevel(root)
    delPop.title('Eliminar todas las filas')
    delPop.resizable(False, False)
    
    eraseAllLbl = Label(delPop, text='Advertencia: Eliminará\ntodas las filas que\naparezcan en pantalla.\n¿desea seguir?', width=35, height=4)
    eraseAllLbl.grid(row=0, column=0)
    
    eraseAllBtn = Button(delPop, text='Hacerlo', width=10, height=2, command=eraseAll)
    eraseAllBtn.grid(row=1, column=0)
            
        
    
# Info part
wdg = Frame(root, bg="#A3E53C", width=400, height=400)
wdg.grid(row=0, column=0, columnspan=3, rowspan=4)
info = Label(wdg, bg="white", fg="black", text="Aqui se mostraran los datos buscados, agregados y quitados")
info.grid(row=0, column=0, columnspan=3)
bandas = Listbox(wdg ,background="#0DD816" ,foreground="#F4173F" ,selectforeground="#1054CC", selectbackground="#3AE1B1", selectborderwidth=3, width=45)
sb = Scrollbar(wdg, orient=VERTICAL)
sb.grid(column=2,sticky='nse')
bandas.config(yscrollcommand=sb.set)
sb.config(command=bandas.yview)
version = Label(root, text='Version 1.1', width=10, height=2, fg='black')
version.grid(row=5, column=3)

# Actions for the data
add = Button(root, text="Agregar", state=DISABLED, width=10, height=2, command=agregar)
add.grid(row=0, column=3)
sub = Button(root, text="Eliminar", state=DISABLED, width=10, height=2, command=quitar)
sub.grid(row=1, column=3)
find = Button(root, text="Encontrar", state=DISABLED, width=10, height=2, command=encontrar)
find.grid(row=2, column=3)

subAll = Button(root, text="Eliminar\nTodo", state=DISABLED, width=10, height=2, command=erasePopup)
subAll.grid(row=1, column=4)

edit = Button(root, text="Editar\nFila", state=DISABLED, width=10, height=2, command=editar)
edit.grid(row=0, column=4)

# Select files
bookinuse = Label(root, text="Eliga su\narchivo", bg="#009009", width=15, height=2)
bookinuse.grid(row=3, column=3)
selectBook = Button(root, text="Elegir \narchivo", width=10, height=2, command=openfile)
selectBook.grid(row=4, column=3)
allpathfile = Label(root, text="")

# Opciones
wdgA = Frame(root, bg="#E84DE5", width=300, height=75)
wdgA.grid(row=4, column=0, rowspan=2, columnspan=3)

medidastxt = Label(wdgA, text="Medidas", width=8, height=1).grid(row=0, column=0)
propiedadestxt = Label(wdgA, text="Propiedades", width=9, height=1).grid(row=0, column=1)
importetxt = Label(wdgA, text="Importe", width=8, height=1).grid(row=0, column=2)

medidas = Entry(wdgA, bg="#F542A9")
medidas.grid(row=1, column=0)
propiedades = Entry(wdgA, bg="#F542A9")
propiedades.grid(row=1, column=1)
importe = Entry(wdgA, bg="#F542A9")
importe.grid(row=1, column=2)

whichsheet = Label(wdgA, text="", width=16, height=2)
whichsheet.grid(row=2, column=0)

# Manipular hojas
selectsheet = ttk.Combobox(wdgA, width=15, height=2)
selectsheet.grid(row=2, column=1)
btnsheet = Button(wdgA, width=10, height=2, command=changeSheet, text="Establecer \nhoja")
btnsheet.grid(row=2, column=2)

nspopup = Button(wdgA, text="Acciones de\nhoja", width=10, height=2, command=showPopup)
nspopup.grid(row=3, column=2)
nspopup.config(state=DISABLED)

root.mainloop()